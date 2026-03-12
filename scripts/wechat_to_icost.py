#!/usr/bin/env python3
"""
wechat_to_icost.py - Convert WeChat Pay bill export (XLSX) to iCost import format.

Usage:
    python3 wechat_to_icost.py <input.xlsx> [output.xlsx]

WeChat export format (row 17 = header, data from row 18):
    交易时间, 交易类型, 交易对方, 商品, 收/支, 金额(元), 支付方式, 当前状态, 交易单号, 商户单号, 备注

iCost import format:
    日期, 类型, 金额, 一级分类, 二级分类, 账户1, 账户2, 备注, 货币, 标签
"""

import sys
import re
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# ---------------------------------------------------------------------------
# Categorization
# ---------------------------------------------------------------------------

TX_TYPE_RULES = [
    ("退款",    "收入", "其他",  "退款"),
    ("红包",    "收入", "其他",  "红包"),
    ("亲属卡",  "支出", "其他",  "亲属卡"),
    ("零钱充值","转账", "其他",  "零钱充值"),
    ("零钱提现","转账", "其他",  "零钱提现"),
    ("转账",    "支出", "转账",  "个人转账"),
]

KEYWORD_RULES = [
    (["餐","饭","面","烤","火锅","奶茶","咖啡","茶饮","luckin","麦当劳","肯德基",
      "包子","粥","饺","牛肉","鱼","西餐","汤包","豆花","米线","烧烤","串","食堂"],
     "餐饮", "三餐"),
    (["骑行","运动","健身","跑步","马拉松","游泳","球","羽毛球","钓鱼"],
     "运动", "运动健身"),
    (["盒马","超市","便利店","名创","京东","拼多多","淘宝","天猫"],
     "购物", "网购/超市"),
    (["滴滴","出行","地铁","公交","高铁","机票","加油","停车","充电桩","充电"],
     "交通", "出行"),
    (["电费","供电","电网","燃气","水费"],
     "居家", "水电燃气"),
    (["物业"],
     "居家", "物业"),
    (["保险"],
     "金融", "保险"),
    (["健康平台","医院","药店","诊所","医疗"],
     "医疗", "医疗健康"),
    (["动物园","探险","宝宝巴士","游乐","摇摇车","亲子"],
     "娱乐", "亲子娱乐"),
    (["美发","理发","美容","美甲","优剪"],
     "个护", "美发美容"),
    (["顺丰","快递","邮政","圆通","中通","韵达"],
     "购物", "快递"),
    (["腾讯云","阿里云","服务器","域名","cdn"],
     "数码", "云服务"),
    (["教育","学习","培训","课程","书店"],
     "教育", "教育培训"),
    (["话费","流量","移动","联通","电信"],
     "通讯", "话费流量"),
]


def categorize(tx_type: str, counterpart: str, product: str):
    t = str(tx_type or "")
    text = (str(counterpart or "") + str(product or "")).lower()

    for keyword, rec_type, cat1, cat2 in TX_TYPE_RULES:
        if keyword in t:
            return rec_type, cat1, cat2

    for keywords, cat1, cat2 in KEYWORD_RULES:
        if any(k.lower() in text for k in keywords):
            return None, cat1, cat2  # rec_type determined later

    return None, "其他", "待分类"


def map_type(income_expense: str, tx_type: str, override_type=None) -> str:
    if override_type:
        return override_type
    ie = str(income_expense or "")
    t = str(tx_type or "")
    if "退款" in t:
        return "收入"
    if ie == "收入":
        return "收入"
    if ie == "支出":
        return "支出"
    return "转账"


def map_account(method: str) -> str:
    m = str(method or "")
    if "信用卡" in m:
        # extract card hint e.g. (8116)
        match = re.search(r'\((\d+)\)', m)
        suffix = f"({match.group(1)})" if match else ""
        return f"信用卡{suffix}"
    if "储蓄卡" in m:
        match = re.search(r'\((\d+)\)', m)
        suffix = f"({match.group(1)})" if match else ""
        return f"储蓄卡{suffix}"
    if "零钱" in m:
        return "微信零钱"
    return "微信"


# ---------------------------------------------------------------------------
# Main conversion
# ---------------------------------------------------------------------------

def convert(input_path: str, output_path: str):
    wb_in = openpyxl.load_workbook(input_path)
    ws_in = wb_in.active

    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active
    ws_out.title = "icost_template"
    ws_out.append(["日期", "类型", "金额", "一级分类", "二级分类", "账户1", "账户2", "备注", "货币", "标签"])

    count = 0
    skipped = 0
    unclassified = []

    for i, row in enumerate(ws_in.iter_rows(values_only=True)):
        # Skip header rows (WeChat export has 16 meta rows + 1 header row)
        if i < 17:
            continue
        if not row[0]:
            continue

        tx_time, tx_type, counterpart, product, income_expense, amount, method, status, *_ = row

        # Parse amount
        amt_str = str(amount or "").replace("¥", "").replace(",", "").strip()
        try:
            amt = float(amt_str)
        except ValueError:
            skipped += 1
            continue

        # Skip zero-amount rows
        if amt == 0:
            skipped += 1
            continue

        # Date
        try:
            dt = datetime.strptime(str(tx_time), "%Y-%m-%d %H:%M:%S")
            date_str = dt.strftime("%Y年%m月%d日 %H:%M:%S")
        except Exception:
            date_str = str(tx_time)

        override_type, cat1, cat2 = categorize(tx_type, counterpart, product)
        rec_type = map_type(income_expense, tx_type, override_type)
        account = map_account(method)

        prod_str = str(product or "")
        note = f"{counterpart} - {prod_str}" if prod_str and prod_str != "/" else str(counterpart or "")

        if cat2 == "待分类":
            unclassified.append((counterpart, product, amt))

        ws_out.append([date_str, rec_type, amt, cat1, cat2, account, None, note, "CNY", None])
        count += 1

    wb_out.save(output_path)

    print(f"✅ Converted {count} records → {output_path}")
    if skipped:
        print(f"   Skipped {skipped} rows (zero amount or parse error)")
    if unclassified:
        print(f"\n⚠️  {len(unclassified)} records need manual categorization (labeled '待分类'):")
        for c, p, a in unclassified[:20]:
            print(f"   ¥{a:.2f}  {c} - {p}")
        if len(unclassified) > 20:
            print(f"   ... and {len(unclassified) - 20} more")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python3 wechat_to_icost.py <input.xlsx> [output.xlsx]")
        sys.exit(1)

    input_file = sys.argv[1]
    output_file = sys.argv[2] if len(sys.argv) > 2 else input_file.replace(".xlsx", "_icost.xlsx")
    convert(input_file, output_file)
